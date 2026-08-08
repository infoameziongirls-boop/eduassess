import os
import sys
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))

from app import app, db
from models import User
from flask_bcrypt import Bcrypt

bcrypt = Bcrypt(app)


def test_import_excel_requires_teacher_or_admin():
    app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)

    with app.app_context():
        db.drop_all()
        db.create_all()

        teacher = User(
            username='perm_teacher',
            password_hash=bcrypt.generate_password_hash('Test@123'),
            role='teacher',
            subject='mathematics',
        )
        student = User(
            username='perm_student',
            password_hash=bcrypt.generate_password_hash('Test@123'),
            role='student',
        )
        db.session.add_all([teacher, student])
        db.session.commit()
        teacher_id = teacher.id
        student_id = student.id

    with app.test_client() as client:
        with client.session_transaction() as sess:
            sess['_user_id'] = str(student_id)
            sess['_fresh'] = True

        response = client.get('/import/excel')
        assert response.status_code == 403

    with app.test_client() as client:
        with client.session_transaction() as sess:
            sess['_user_id'] = str(teacher_id)
            sess['_fresh'] = True

        response = client.get('/import/excel')
        assert response.status_code == 200


def test_bulk_roster_form_renders_subject_options_from_config():
    app.config.update(
        TESTING=True,
        WTF_CSRF_ENABLED=False,
        STUDY_AREAS=[('Science', 'Science')],
        STUDY_AREA_SUBJECTS={'Science': {'core': ['Biology'], 'electives': ['Chemistry']}},
    )

    with app.app_context():
        db.drop_all()
        db.create_all()

        teacher = User(
            username='roster_teacher',
            password_hash=bcrypt.generate_password_hash('Test@123'),
            role='teacher',
            subject='biology',
        )
        db.session.add(teacher)
        db.session.commit()
        teacher_id = teacher.id

    with app.test_client() as client:
        with client.session_transaction() as sess:
            sess['_user_id'] = str(teacher_id)
            sess['_fresh'] = True

        response = client.get('/assessments/bulk_roster')
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert 'name="subject"' in html
        assert 'Biology' in html
        assert 'Chemistry' in html


def test_prefilled_roster_workbook_unlocks_score_and_comments_cells(tmp_path):
    from template_updater import create_prefilled_roster_template
    from models import Student

    with app.app_context():
        db.drop_all()
        db.create_all()
        student = Student(
            student_number='S001',
            first_name='Ada',
            last_name='Lovelace',
            class_name='Form 1',
            study_area='Science',
            reference_number='R001',
        )
        db.session.add(student)
        db.session.commit()

        output_path = tmp_path / 'roster.xlsx'
        create_prefilled_roster_template(
            str(output_path),
            [student],
            subject='Biology',
            class_name='Form 1',
            term='Term 1',
            academic_year='2026',
            session='Morning',
            category='ica1',
            assessor='Test Assessor',
        )

        wb = load_workbook(output_path)
        ws = wb['Roster']
        assert ws['E2'].protection.locked is False
        assert ws['J2'].protection.locked is False
        assert ws['A3'].value is not None
        assert 'Editable' in str(ws['A3'].value)
        assert ws['A12'].value is None
        assert ws['E2'].fill.fgColor.rgb != ws['A2'].fill.fgColor.rgb
        assert ws['J2'].fill.fgColor.rgb != ws['A2'].fill.fgColor.rgb


def test_prefilled_roster_workbook_places_note_after_all_students(tmp_path):
    from template_updater import create_prefilled_roster_template
    from models import Student

    with app.app_context():
        db.drop_all()
        db.create_all()

        students = []
        for idx in range(1, 12):
            students.append(Student(
                student_number=f'S{idx:03d}',
                first_name=f'First{idx}',
                last_name=f'Last{idx}',
                class_name='Form 1',
                study_area='Science',
                reference_number=f'R{idx:03d}',
            ))
        db.session.add_all(students)
        db.session.commit()

        output_path = tmp_path / 'roster_11.xlsx'
        create_prefilled_roster_template(
            str(output_path),
            students,
            subject='Biology',
            class_name='Form 1',
            term='Term 1',
            academic_year='2026',
            session='Morning',
            category='ica1',
            assessor='Test Assessor',
        )

        wb = load_workbook(output_path)
        ws = wb['Roster']
        assert ws['A13'].value is not None
        assert 'Editable' in str(ws['A13'].value)
        assert ws['A12'].value != ws['A13'].value
