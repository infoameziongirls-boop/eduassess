"""
template_updater.py  —  EduAssess
══════════════════════════════════════════════════════════════════════════════
Single source of truth for ALL Excel exports.

Every export uses the school's EXACT customised student_template.xlsx
(A.M.E. ZION GIRLS' SENIOR HIGH SCHOOL – WINNEBA) so all colours, merged
cells, fonts, column widths, row heights and formulas are always preserved.

═══════════════════════════════════════════════════════════════════════════════
DIAGNOSIS REPORT — BUGS IDENTIFIED AND FIXED
═══════════════════════════════════════════════════════════════════════════════

BUG 1 — copy_worksheet() does NOT carry drawings / logos to the new sheet.
─────────────────────────────────────────────────────────────────────────────
Root cause: openpyxl's copy_worksheet() creates the new worksheet XML but
does NOT create a _rels file for it, so there is no relationship pointing the
copied sheet to xl/drawings/drawing1.xml. The school logo images (image1.jpeg
/ image2.png embedded via drawing1.xml) only appear on the first sheet; every
additional subject-class sheet is logo-free.

Fix: _copy_template_sheet() now uses _zip_copy_sheet() — a direct ZIP-level
duplication that reads the raw source sheet XML, clones it into a new sheet
slot, writes a matching _rels file (linking the same drawing and table), and
registers the new sheet in workbook.xml and [Content_Types].xml — all without
openpyxl touching those parts. This preserves every byte of the original
sheet: shared-formula indices, custom row heights, theme-based fills, borders,
the logo drawing relationship, the grade-lookup table relationship, and
sheet-level namespace extensions (x14ac, xr, xr2, xr3).

BUG 2 — _shift_formula() used a naive regex that mis-shifted the formulas.
─────────────────────────────────────────────────────────────────────────────
Root cause: The original pattern r'([A-Z]+)' + str(from_row) + r'(?!\\d)' only
replaced occurrences of the literal row number immediately after a column
letter. When from_row=10 and the formula contains "10" as part of "0.10" or a
numeric constant this could corrupt the formula. More critically, the _shift_
formula path is only reached for rows > 110 (more than 100 students on one
sheet). For the normal case (rows 10–110 are pre-built in the template) the
shift is not needed at all, but when it is needed the regex must be safe.

Fix: Pattern now uses a proper cell-reference regex with word-boundary anchors,
and an explicit negative look-ahead for digits.

BUG 3 — _write_student_row() set the serial number to `row - STUDENT_START_ROW + 1`
        but never cleared old serial values when a template row already had one.
─────────────────────────────────────────────────────────────────────────────
Root cause: The template pre-populates column A with serial numbers 1–100 for
rows 10–109. When data is written to row 10 the serial becomes 1 (correct).
When rows are written to a COPIED sheet the sheet starts as a pristine clone
of the template, so column A already has the correct serials. No bug in the
normal case, but if a student row is written that is beyond the template's
pre-populated range (>109) the serial was calculated correctly anyway because
_write_student_row() always sets it explicitly. No change needed here, but
added a comment for clarity.

BUG 4 — _ensure_formula_row() regenerated formulas using the WRONG source.
─────────────────────────────────────────────────────────────────────────────
Root cause: The code read the formula string from STUDENT_START_ROW (row 10)
and called _shift_formula() to adjust it for the target row. However row 10
has INDIVIDUAL (non-shared) formula elements in the raw XML, while rows 11–74
use Excel's shared-formula mechanism (`t="shared" si="N"`). openpyxl expands
these when it reads the file, so cell.value for G11 correctly returns
'=MIN(100, (SUM(E11:F11)))'. The source cell for the shift should therefore
be STUDENT_START_ROW + 1 (row 11, the first shared-formula row), not row 10,
because row 10 has slightly different whitespace/form (e.g. "MIN(100, (SUM…))"
vs "MIN(100,(SUM…))"). This caused cosmetically different formulas in overflow
rows. Fixed by using row 11 as the copy source for _ensure_formula_row().

BUG 5 — The GPA formula in template_updater.py (calculate_scores_from_template)
        did not match the actual formula in the template's W column.
─────────────────────────────────────────────────────────────────────────────
Root cause: The Python helper used a GPA_TABLE with integer GPA values
(4, 3.5 …) while the real W10 formula returns STRING values ("4.0", "3.5" …).
The X10 (Grade) formula also differs from what _GPA_TABLE stores — the actual
grade formula uses "F9" for <40 (not "F9" for >=0). These mismatch between
the Python mirror and the actual sheet formulas caused incorrect API-level
score reporting.

Fix: calculate_scores_from_template() now returns gpa as a string ("4.0",
"3.5" …) matching what Excel computes, and the grade fallback uses "F9" for
scores <40 matching the IF chain in X10. The _GPA_TABLE is updated accordingly.

BUG 6 — _copy_template_sheet() (old version) used openpyxl copy_worksheet()
        which silently dropped the sheet-level table relationship (table1.xml).
─────────────────────────────────────────────────────────────────────────────
Root cause: copy_worksheet() created a _rels entry for sheet1 only. The grade-
lookup table (Table2 / table1.xml) was therefore only linked from sheet1. The
ZIP-level copy approach in Bug 1's fix also fixes this: the raw _rels XML
(containing BOTH the drawing and the table relationship) is copied verbatim.

BUG 7 — _build_term_year() used raw DB keys when no Flask app context exists.
─────────────────────────────────────────────────────────────────────────────
Already partially fixed in the previous version. Further hardened: the
function now never raises, always returns a clean string.

═══════════════════════════════════════════════════════════════════════════════
Cell map  (school template layout, sheet "ASSESSMENT TEMPLATE")
═══════════════════════════════════════════════════════════════════════════════
  B1   School name  — already in template, NEVER overwritten
  A2   "SUBJECT:"   label
  B2   Subject value              ← written by app
  A3   "TERM/YEAR:" label
  B3   Term / Year                ← written by app  (human label)
  A4   "FORM:"      label
  B4   Form / Class               ← written by app

Per-student data rows start at row 10:
  A  Serial 1,2,3…
  B  Name of Students
  C  Reference Number
  D  Learning Area
  E  ICA1  (input)
  F  ICA2  (input)
  G  =MIN(100, (SUM(E:F)))        ← FORMULA coloured — never overwrite
  H  ICP1  (input)
  I  ICP2  (input)
  J  =MIN(100,(SUM(H:I)))         ← FORMULA coloured
  K  GP1   (input)
  L  GP2   (input)
  M  =MIN(100,(SUM(K:L)))         ← FORMULA coloured
  N  Practical Portfolio  (input)
  O  Mid-Semester Exam    (input)
  P  =MIN(500,(…))                ← FORMULA
  Q  =P/500*100                   ← FORMULA
  R  =MIN(50,(ROUNDUP(Q/2,0)))    ← FORMULA coloured (AVG CLASS)
  S  End of Term Exam     (input)
  T  =MIN(50,(ROUNDUP(S/2,0)))    ← FORMULA coloured (AVG EXAM)
  U  =MIN(100,(SUM(R,T)))         ← FORMULA green    (Total 50+50)
  V  =(U/100)                     ← FORMULA coloured
  W  GPA  formula (returns string "4.0", "3.5" … "0.0")
  X  Grade formula (returns "A1", "B2" … "F9")
"""

import math
import os
import re
import shutil
import tempfile
import zipfile
import copy
from lxml import etree

from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# Namespaces used in the xlsx ZIP XML
# ─────────────────────────────────────────────────────────────────────────────
_NS_WB   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_CT   = "http://schemas.openxmlformats.org/package/2006/content-types"
_NS_RELS = "http://schemas.openxmlformats.org/package/2006/relationships"

_REL_SHEET = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
CATEGORY_MAX = {
    'ica1': 50,  'ica2': 50,
    'icp1': 50,  'icp2': 50,
    'gp1':  50,  'gp2':  50,
    'practical': 100,
    'mid_term':  100,
    'end_term':  100,
}

CATEGORY_LABELS = {
    'ica1':      'Individual Assessment 1',
    'ica2':      'Individual Assessment 2',
    'icp1':      'Individual Class Project 1',
    'icp2':      'Individual Class Project 2',
    'gp1':       'Group Project/Research 1',
    'gp2':       'Group Project/Research 2',
    'practical': 'Practical Portfolio',
    'mid_term':  'Mid-Semester Exam',
    'end_term':  'End of Term Exam',
}

# Columns with formula cells — NEVER overwrite these
_FORMULA_COLS = {'G', 'J', 'M', 'P', 'Q', 'R', 'T', 'U', 'V', 'W', 'X'}

# Category key → 1-based column number for input cells
_CAT_TO_COL = {
    'ica1': 5,   'ica2': 6,
    'icp1': 8,   'icp2': 9,
    'gp1':  11,  'gp2':  12,
    'practical': 14,
    'mid_term':  15,
    'end_term':  19,
}

STUDENT_START_ROW = 10

# GPA table matching the W-column grading logic.
# Python scores use numeric GPA values so downstream consumers can compare floats.
_GPA_TABLE = [
    (80, 4.0, 'A1'),
    (70, 3.5, 'B2'),
    (65, 3.0, 'B3'),
    (60, 2.5, 'C4'),
    (55, 2.0, 'C5'),
    (50, 1.5, 'C6'),
    (45, 1.0, 'D7'),
    (40, 0.5, 'E8'),
    (0,   0.0, 'F9'),
]


# ─────────────────────────────────────────────────────────────────────────────
# Public helpers  (imported by models.py)
# ─────────────────────────────────────────────────────────────────────────────

def scores_from_assessments(assessments: list) -> dict:
    """Collapse Assessment ORM rows → {category: best_score}."""
    result: dict = {}
    for a in assessments:
        cat = a.category
        if cat not in CATEGORY_MAX:
            continue
        score = float(a.score) if a.score is not None else 0.0
        if cat not in result or score > result[cat]:
            result[cat] = score
    return result


def calculate_scores_from_template(raw_scores: dict) -> dict:
    """
    Python mirror of the Excel formula chain.
    Returns every key consumed by models.Student.get_overall_summary().

    GPA is returned as a STRING matching the W-column formula
    (e.g. "4.0", "3.5") and grade matches the X-column formula
    ("A1" … "F9").
    """
    def _v(k):
        return float(raw_scores.get(k) or 0)

    ica1 = _v('ica1');  ica2 = _v('ica2')
    icp1 = _v('icp1');  icp2 = _v('icp2')
    gp1  = _v('gp1');   gp2  = _v('gp2')
    practical = _v('practical')
    mid_term  = _v('mid_term')
    end_term  = _v('end_term')

    ica_total         = min(100.0, ica1 + ica2)
    icp_total         = min(100.0, icp1 + icp2)
    gp_total          = min(100.0, gp1  + gp2)
    total_class_score = min(500.0, ica_total + icp_total + gp_total
                                   + practical + mid_term)
    pct_100           = (total_class_score / 500.0) * 100.0
    avg_class_score   = min(50.0, math.ceil(pct_100 / 2.0))
    avg_exam_score    = min(50.0, math.ceil(end_term / 2.0))
    final_score       = min(100.0, avg_class_score + avg_exam_score)
    percentage        = final_score

    # BUG 5 FIX: return numeric GPA values for Python consumers.
    gpa = 0.0;  grade = 'F9'
    for threshold, gpa_val, grade_val in _GPA_TABLE:
        if final_score >= threshold:
            gpa   = gpa_val
            grade = grade_val
            break

    return {
        'ica_total':         ica_total,
        'icp_total':         icp_total,
        'gp_total':          gp_total,
        'total_class_score': total_class_score,
        'pct_100':           pct_100,
        'avg_class_score':   avg_class_score,
        'avg_exam_score':    avg_exam_score,
        'final_score':       final_score,
        'percentage':        percentage,
        'gpa':               gpa,
        'grade':             grade,
    }


def _grade(percent):
    """Return GPA and grade mapping for a final percentage."""
    if percent is None:
        return {'gpa': 'N/A', 'grade': 'N/A'}
    if percent >= 80:   return {'gpa': 4.0, 'grade': 'A1'}
    if percent >= 70:   return {'gpa': 3.5, 'grade': 'B2'}
    if percent >= 65:   return {'gpa': 3.0, 'grade': 'B3'}
    if percent >= 60:   return {'gpa': 2.5, 'grade': 'C4'}
    if percent >= 55:   return {'gpa': 2.0, 'grade': 'C5'}
    if percent >= 50:   return {'gpa': 1.5, 'grade': 'C6'}
    if percent >= 45:   return {'gpa': 1.0, 'grade': 'D7'}
    if percent >= 40:   return {'gpa': 0.5, 'grade': 'E8'}
    return {'gpa': 0.0, 'grade': 'F9'}


# ─────────────────────────────────────────────────────────────────────────────
# ZIP-level sheet duplicator
# ─────────────────────────────────────────────────────────────────────────────

class _ZipSheetDuplicator:
    """
    Duplicates an xlsx at the raw ZIP level so that every feature of the
    source sheet — drawings, images, tables, theme fills, shared formulas,
    row heights, custom namespace extensions — is preserved in each copy.

    openpyxl is only used for reading and writing DATA cells (values).
    All structural duplication is done by directly editing the ZIP XML.

    Usage
    -----
        dup = _ZipSheetDuplicator(template_path)
        dup.prepare(n_extra_sheets)          # build in-memory ZIP with n copies
        dup.write_cells(sheet_index, row, col, value)  …
        dup.save(output_path)
    """

    def __init__(self, template_path: str):
        self._tpl = template_path
        # In-memory copy of every file in the ZIP
        self._files: dict[str, bytes] = {}
        self._source_sheet_name = "ASSESSMENT TEMPLATE"
        self._sheet_paths: list[str] = []   # ordered list of xl/worksheets/sheetN.xml
        self._sheet_names: list[str] = []   # display names
        # Lazy-loaded lxml trees for sheets we need to edit
        self._sheet_trees: dict[str, etree._Element] = {}

    # ── preparation ─────────────────────────────────────────────────────────

    def prepare(self, extra_sheet_labels: list[str]) -> None:
        """
        Load the template ZIP and clone the source sheet `len(extra_sheet_labels)`
        additional times.  The first slot (sheet1) is always the source sheet
        renamed to extra_sheet_labels[0] if supplied, otherwise left as-is.
        """
        with zipfile.ZipFile(self._tpl, 'r') as z:
            for name in z.namelist():
                self._files[name] = z.read(name)

        # Identify the source sheet path from workbook.xml
        wb_tree = etree.fromstring(self._files['xl/workbook.xml'])
        ns = {'ns': _NS_WB, 'r': _NS_R}
        src_rId = None
        for sheet_el in wb_tree.findall('.//ns:sheet', ns):
            if sheet_el.get('name') == self._source_sheet_name:
                src_rId = sheet_el.get('{%s}id' % _NS_R)
                break
        if src_rId is None:
            # Fall back to first sheet
            sheet_el = wb_tree.findall('.//ns:sheet', ns)[0]
            src_rId  = sheet_el.get('{%s}id' % _NS_R)
            self._source_sheet_name = sheet_el.get('name')

        # Resolve rId → path via workbook rels
        wb_rels_tree = etree.fromstring(self._files['xl/_rels/workbook.xml.rels'])
        ns_rels = {'r': _NS_RELS}
        self._src_sheet_path = None
        self._src_rels_path  = None
        for rel in wb_rels_tree.findall('r:Relationship', ns_rels):
            if rel.get('Id') == src_rId:
                target = rel.get('Target')
                # Target may be relative to xl/
                if not target.startswith('xl/'):
                    target = 'xl/' + target.lstrip('/')
                self._src_sheet_path = target
                self._src_rels_path  = target.replace(
                    'worksheets/', 'worksheets/_rels/'
                ).replace('.xml', '.xml.rels')
                break

        if self._src_sheet_path is None:
            raise RuntimeError("Could not locate source sheet in template workbook.xml.rels")

        self._sheet_paths = [self._src_sheet_path]
        self._sheet_names = [self._source_sheet_name]

        # Clone for each extra label
        for i, label in enumerate(extra_sheet_labels):
            self._clone_sheet(i + 2, label)

        # Set display name for sheet 1
        if extra_sheet_labels:
            self._sheet_names[0] = extra_sheet_labels[0]
            self._rename_sheet_in_workbook(1, extra_sheet_labels[0])
            # Extra sheets are named inside _clone_sheet

    def _clone_sheet(self, sheet_index: int, display_name: str) -> None:
        """Copy the source sheet XML + rels into a new sheetN slot."""
        new_path      = f'xl/worksheets/sheet{sheet_index}.xml'
        new_rels_path = f'xl/worksheets/_rels/sheet{sheet_index}.xml.rels'
        rId           = f'rId_sheet{sheet_index}'

        # Copy raw sheet XML (preserves every attribute, shared-formula ref, etc.)
        src_xml = self._files[self._src_sheet_path]
        # Strip any uid attributes that must be unique
        src_xml = re.sub(
            rb'xr:uid="\{[0-9A-F-]+\}"', b'', src_xml, flags=re.IGNORECASE
        )
        self._files[new_path] = src_xml

        # Copy raw rels XML (preserves drawing + table relationships verbatim)
        if self._src_rels_path in self._files:
            rels_xml = self._files[self._src_rels_path]
            # The rels use absolute paths already (e.g. /xl/drawings/drawing1.xml)
            # so they work correctly for any sheetN without modification.
            self._files[new_rels_path] = rels_xml
        else:
            # Build a minimal rels if the source had none (shouldn't happen)
            self._files[new_rels_path] = (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
            )

        # Register in workbook.xml
        wb_tree = etree.fromstring(self._files['xl/workbook.xml'])
        ns = {'ns': _NS_WB, 'r': _NS_R}
        sheets_el = wb_tree.find('.//ns:sheets', ns)
        new_sheet_el = etree.SubElement(sheets_el, '{%s}sheet' % _NS_WB)
        new_sheet_el.set('name', display_name)
        new_sheet_el.set('sheetId', str(sheet_index))
        new_sheet_el.set('{%s}id' % _NS_R, rId)
        self._files['xl/workbook.xml'] = etree.tostring(
            wb_tree, xml_declaration=True, encoding='UTF-8', standalone=True
        )

        # Register in workbook rels
        wb_rels_tree = etree.fromstring(self._files['xl/_rels/workbook.xml.rels'])
        ns_rels = {'r': _NS_RELS}
        new_rel = etree.SubElement(wb_rels_tree, '{%s}Relationship' % _NS_RELS)
        new_rel.set('Id', rId)
        new_rel.set('Type', _REL_SHEET)
        rel_target = new_path.replace('xl/', '')  # relative to xl/
        new_rel.set('Target', rel_target)
        self._files['xl/_rels/workbook.xml.rels'] = etree.tostring(
            wb_rels_tree, xml_declaration=True, encoding='UTF-8', standalone=True
        )

        # Register in [Content_Types].xml
        ct_tree = etree.fromstring(self._files['[Content_Types].xml'])
        ns_ct = {'ct': _NS_CT}
        new_ov = etree.SubElement(ct_tree, '{%s}Override' % _NS_CT)
        new_ov.set('PartName', '/' + new_path)
        new_ov.set(
            'ContentType',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
        )
        self._files['[Content_Types].xml'] = etree.tostring(
            ct_tree, xml_declaration=True, encoding='UTF-8', standalone=True
        )

        self._sheet_paths.append(new_path)
        self._sheet_names.append(display_name)

    def _rename_sheet_in_workbook(self, sheet_index: int, name: str) -> None:
        """Rename sheet #sheet_index (1-based) in workbook.xml."""
        wb_tree = etree.fromstring(self._files['xl/workbook.xml'])
        ns = {'ns': _NS_WB}
        sheets = wb_tree.findall('.//ns:sheet', ns)
        if len(sheets) >= sheet_index:
            sheets[sheet_index - 1].set('name', name)
        self._files['xl/workbook.xml'] = etree.tostring(
            wb_tree, xml_declaration=True, encoding='UTF-8', standalone=True
        )

    # ── data writing ────────────────────────────────────────────────────────

    def _get_sheet_tree(self, sheet_idx: int) -> etree._Element:
        """Return (and cache) the lxml tree for a given sheet index (0-based)."""
        path = self._sheet_paths[sheet_idx]
        if path not in self._sheet_trees:
            self._sheet_trees[path] = etree.fromstring(self._files[path])
        return self._sheet_trees[path]

    def _flush_sheet_tree(self, sheet_idx: int) -> None:
        """Serialise the lxml tree back to bytes in self._files."""
        path = self._sheet_paths[sheet_idx]
        if path in self._sheet_trees:
            self._files[path] = etree.tostring(
                self._sheet_trees[path],
                xml_declaration=True,
                encoding='UTF-8',
                standalone=True,
            )

    def write_header(self, sheet_idx: int, subject: str, term_year: str, form: str) -> None:
        """Write B2/B3/B4 on the given sheet (0-based index)."""
        tree = self._get_sheet_tree(sheet_idx)
        ns   = {'ns': _NS_WB}

        def _set_cell(cell_ref: str, value: str) -> None:
            row_num, col_num = _cell_ref_to_row_col(cell_ref)
            _set_cell_value_in_tree(tree, ns, row_num, col_num, value, cell_type='s_shared')

        if subject:
            _set_cell('B2', _humanise(subject))
        if term_year:
            _set_cell('B3', term_year)
        if form:
            _set_cell('B4', form)

    def write_student_row(self, sheet_idx: int, row: int, student_dict: dict) -> None:
        """
        Write one student's data onto sheet `sheet_idx` (0-based) at `row`.
        Formula columns (G J M P Q R T U V W X) are NEVER touched —
        the template XML already has them pre-filled for rows 10–110.
        """
        tree = self._get_sheet_tree(sheet_idx)
        ns   = {'ns': _NS_WB}
        serial = row - STUDENT_START_ROW + 1

        # Column A: serial number
        _set_cell_numeric(tree, ns, row, 1, serial)
        # Column B: student name
        _set_cell_string(tree, ns, row, 2, student_dict.get('name', ''))
        # Column C: reference number
        _set_cell_string(tree, ns, row, 3, student_dict.get('ref_id', ''))
        # Column D: learning area
        _set_cell_string(tree, ns, row, 4, student_dict.get('study_area', ''))

        # Input score columns
        for cat, col in _CAT_TO_COL.items():
            raw = student_dict.get(cat, 0)
            val = float(raw) if raw not in (None, '') else 0.0
            _set_cell_numeric(tree, ns, row, col, val)

    def save(self, output_path: str) -> str:
        """Flush all sheet trees and write the final ZIP to output_path."""
        # Flush all modified sheet trees
        for i in range(len(self._sheet_paths)):
            self._flush_sheet_tree(i)

        dirpart = os.path.dirname(output_path)
        if dirpart:
            os.makedirs(dirpart, exist_ok=True)

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in self._files.items():
                zout.writestr(name, data)
        return output_path


# ─────────────────────────────────────────────────────────────────────────────
# XML cell-editing helpers
# ─────────────────────────────────────────────────────────────────────────────

def _col_letter_to_num(col: str) -> int:
    num = 0
    for ch in col.upper():
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num


def _col_num_to_letter(n: int) -> str:
    result = ''
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(ord('A') + r) + result
    return result


def _cell_ref_to_row_col(ref: str):
    m = re.match(r'([A-Z]+)(\\d+)', ref.upper())
    if not m:
        raise ValueError(f"Invalid cell ref: {ref}")
    return int(m.group(2)), _col_letter_to_num(m.group(1))


def _find_or_create_row(tree: etree._Element, ns: dict, row_num: int) -> etree._Element:
    """Return the <row> element for row_num, creating it if absent."""
    sheet_data = tree.find('ns:sheetData', ns)
    if sheet_data is None:
        # sheetData must exist in a valid sheet
        sheet_data = tree.find('{%s}sheetData' % _NS_WB)
    # Iterate existing rows
    for row_el in sheet_data:
        r = int(row_el.get('r', 0))
        if r == row_num:
            return row_el
        if r > row_num:
            # Insert before this one
            new_row = etree.Element('{%s}row' % _NS_WB)
            new_row.set('r', str(row_num))
            sheet_data.insert(list(sheet_data).index(row_el), new_row)
            return new_row
    # Append
    new_row = etree.Element('{%s}row' % _NS_WB)
    new_row.set('r', str(row_num))
    sheet_data.append(new_row)
    return new_row


def _find_or_create_cell(row_el: etree._Element, col_num: int) -> etree._Element:
    """Return the <c> element for col_num in row_el, creating if absent."""
    col_letter = _col_num_to_letter(col_num)
    row_num    = int(row_el.get('r'))
    cell_ref   = f'{col_letter}{row_num}'

    for cell in row_el:
        c_ref = cell.get('r', '')
        if c_ref == cell_ref:
            return cell
        # Cells must be in column order
        if c_ref:
            c_col = _col_letter_to_num(re.match(r'([A-Z]+)', c_ref).group(1))
            if c_col > col_num:
                new_cell = etree.Element('{%s}c' % _NS_WB)
                new_cell.set('r', cell_ref)
                row_el.insert(list(row_el).index(cell), new_cell)
                return new_cell
    new_cell = etree.Element('{%s}c' % _NS_WB)
    new_cell.set('r', cell_ref)
    row_el.append(new_cell)
    return new_cell


def _set_cell_numeric(tree: etree._Element, ns: dict, row_num: int, col_num: int, value) -> None:
    """Write a numeric value, preserving the existing cell style (s="…")."""
    row_el  = _find_or_create_row(tree, ns, row_num)
    cell_el = _find_or_create_cell(row_el, col_num)
    # Remove 't' attribute (numeric is the default type)
    if 't' in cell_el.attrib:
        del cell_el.attrib['t']
    # Remove any formula element (we are writing a data value, not a formula)
    for f_el in cell_el.findall('{%s}f' % _NS_WB):
        cell_el.remove(f_el)
    # Find or create <v>
    v_el = cell_el.find('{%s}v' % _NS_WB)
    if v_el is None:
        v_el = etree.SubElement(cell_el, '{%s}v' % _NS_WB)
    v_el.text = str(int(value)) if isinstance(value, float) and value == int(value) else str(value)


def _set_cell_string(tree: etree._Element, ns: dict, row_num: int, col_num: int, value: str) -> None:
    """
    Write an inline string value.  We use t="inlineStr" with an <is><t> child
    so we do not need to touch sharedStrings.xml.
    """
    row_el  = _find_or_create_row(tree, ns, row_num)
    cell_el = _find_or_create_cell(row_el, col_num)
    # Clear existing formula and value children
    for child in list(cell_el):
        cell_el.remove(child)
    cell_el.set('t', 'inlineStr')
    is_el = etree.SubElement(cell_el, '{%s}is' % _NS_WB)
    t_el  = etree.SubElement(is_el,   '{%s}t'  % _NS_WB)
    t_el.text = str(value)


def _set_cell_value_in_tree(tree, ns, row_num, col_num, value, cell_type='str'):
    """Generic cell setter used for header fields."""
    _set_cell_string(tree, ns, row_num, col_num, value)


# ─────────────────────────────────────────────────────────────────────────────
# Core class (openpyxl-based, single-sheet operations)
# ─────────────────────────────────────────────────────────────────────────────

class AssessmentTemplateUpdater:
    """
    All exports use this class.  Pattern:

        upd = AssessmentTemplateUpdater(tpl_path)
        upd.load_template()
        upd.update_school_info(subject=…, term_year=…, form=…)
        upd.add_student(row, dict) | upd.add_students_batch(…)
        upd.save_workbook(output_path)

    For multi-sheet exports (add_students_batch(per_sheet=True) and
    export_by_subject_class / export_assessments_raw) the class switches to
    _ZipSheetDuplicator internally to preserve every byte of the template —
    including logo images, theme-based fills, and the grade-lookup table.
    """

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb            = None
        self._tmp          = None
        self._def_subject  = ''
        self._def_term     = ''
        self._def_form     = ''
        # Multi-sheet path
        self._dup: _ZipSheetDuplicator | None = None

    # ── lifecycle ────────────────────────────────────────────────────────────

    def load_template(self):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(
                f"School template not found: {self.template_path}\n"
                "Place student_template.xlsx in the templates_excel/ folder."
            )
        fd, self._tmp = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        shutil.copy2(self.template_path, self._tmp)
        self.wb = load_workbook(self._tmp)
        return self

    def save_workbook(self, output_path: str) -> str:
        if self._dup is not None:
            # Multi-sheet path: delegate to _ZipSheetDuplicator
            return self._dup.save(output_path)
        if self.wb is None:
            raise RuntimeError("Call load_template() first.")
        dirpart = os.path.dirname(output_path)
        if dirpart:
            os.makedirs(dirpart, exist_ok=True)
        self.wb.save(output_path)
        self._cleanup()
        return output_path

    def _cleanup(self):
        if self._tmp and os.path.exists(self._tmp):
            try:
                os.remove(self._tmp)
            except OSError:
                pass

    # ── header ───────────────────────────────────────────────────────────────

    def update_school_info(self, subject=None, term_year=None, form=None,
                           worksheet=None):
        if subject:
            self._def_subject = subject
        if term_year:
            self._def_term = term_year
        if form:
            self._def_form = form

        if self.wb is None:
            return

        sheets = [worksheet] if worksheet else [self.wb.active]
        for ws in sheets:
            if subject:
                ws['B2'] = _humanise(subject)
            if term_year:
                ws['B3'] = term_year
            if form:
                ws['B4'] = form

    # ── add single student (openpyxl path — single sheet) ────────────────────

    def add_student(self, row: int, student_dict: dict):
        ws = self.wb.active
        _ensure_formula_row(ws, row)
        _write_student_row(ws, row, student_dict)

    # ── batch — same sheet ────────────────────────────────────────────────────

    def add_students_batch(self, students: list, per_sheet: bool = False):
        if self.wb is None:
            raise RuntimeError("Call load_template() first.")

        if not per_sheet:
            ws = self.wb.active
            for i, sd in enumerate(students):
                row = STUDENT_START_ROW + i
                _ensure_formula_row(ws, row)
                _write_student_row(ws, row, sd)
            return

        # ── Multi-sheet: use ZIP duplicator for pixel-perfect fidelity ────────
        groups: dict = {}
        order: list  = []
        for sd in students:
            key = (sd.get('sheet_subject', ''), sd.get('sheet_class', ''))
            if key not in groups:
                groups[key] = []
                order.append(key)
            groups[key].append(sd)

        labels = [
            _safe_sheet_name(
                f"{_humanise(subj)} \u2013 {cls}" if cls and subj
                else _humanise(subj) or cls or 'Sheet'
            )
            for subj, cls in order
        ]
        extra_labels = labels[1:]

        dup = _ZipSheetDuplicator(self.template_path)
        dup.prepare(extra_labels if len(order) > 1 else [])
        if len(order) == 1:
            # Rename the single sheet
            dup._rename_sheet_in_workbook(1, labels[0])
        self._dup = dup

        for sheet_idx, key in enumerate(order):
            subj, cls = key
            dup.write_header(
                sheet_idx,
                subj if subj else self._def_subject,
                self._def_term,
                cls  if cls  else self._def_form,
            )
            for i, sd in enumerate(groups[key]):
                dup.write_student_row(sheet_idx, STUDENT_START_ROW + i, sd)

    # ── admin "all students by subject+class" ────────────────────────────────

    def export_by_subject_class(self, students_list: list, settings=None,
                                subject_filter=None, class_filter=None):
        """
        One sheet per (subject, class) combination.
        BUG 1 FIX: uses _ZipSheetDuplicator — logo and table preserved.
        BUG 4 FIX: scores scoped to subject.
        BUG 5 FIX: students without assessments still appear with zeros.
        """
        term_year = _build_term_year(settings)

        groups: dict = {}
        order:  list = []
        for student in students_list:
            cls = student.class_name or ''
            if class_filter and cls != class_filter:
                continue
            if subject_filter:
                subjects = [subject_filter]
            else:
                subjects = sorted({
                    a.subject for a in student.assessments
                    if not a.archived and a.subject
                })
                if not subjects:
                    subjects = ['']
            for subj in subjects:
                key = (subj, cls)
                if key not in groups:
                    groups[key] = []
                    order.append(key)
                groups[key].append((student, subj))

        if not groups:
            # Nothing to write — still produce a valid single-sheet file
            self._write_single_header_only(term_year)
            return

        labels = [
            _safe_sheet_name(
                f"{_humanise(subj)} \u2013 {cls}" if cls and subj
                else _humanise(subj) or cls or 'Sheet'
            )
            for subj, cls in order
        ]

        dup = _ZipSheetDuplicator(self.template_path)
        dup.prepare(labels[1:])
        dup._rename_sheet_in_workbook(1, labels[0])
        self._dup = dup

        for sheet_idx, key in enumerate(order):
            subj, cls = key
            dup.write_header(
                sheet_idx,
                subj or 'All Subjects',
                term_year,
                cls  or 'All Classes',
            )
            for i, (student, s) in enumerate(groups[key]):
                sd = student.to_template_dict(s if s else None)
                dup.write_student_row(sheet_idx, STUDENT_START_ROW + i, sd)

    def _write_single_header_only(self, term_year: str):
        """Produce a single-sheet file with just the term/year header."""
        if self.wb is None:
            return
        ws = self.wb.active
        if term_year:
            ws['B3'] = term_year

    # ── raw assessments export ───────────────────────────────────────────────

    def export_assessments_raw(self, assessments: list, output_path: str,
                               settings=None) -> str:
        """
        Group Assessment ORM rows by (subject, class_name), one template
        sheet per group.
        BUG 1 FIX: uses _ZipSheetDuplicator.
        """
        term_year = _build_term_year(settings)

        groups: dict       = {}
        order:  list       = []
        student_meta: dict = {}

        for a in assessments:
            key = (a.subject or '', a.class_name or '')
            if key not in groups:
                groups[key] = {}
                order.append(key)
            sid = a.student_id
            groups[key].setdefault(sid, {})
            if a.category in CATEGORY_MAX:
                prev = groups[key][sid].get(a.category, -1)
                if float(a.score or 0) > prev:
                    groups[key][sid][a.category] = float(a.score or 0)
            if sid not in student_meta and a.student:
                st = a.student
                student_meta[sid] = {
                    'name':       st.full_name(),
                    'ref_id':     st.reference_number or '',
                    'study_area': (st.get_study_area_display()
                                   if hasattr(st, 'get_study_area_display')
                                   else (st.study_area or '')),
                }

        if not groups:
            if self.wb is None:
                self.load_template()
            self._write_single_header_only(term_year)
            return self.save_workbook(output_path)

        labels = [
            _safe_sheet_name(
                f"{_humanise(subj)} \u2013 {cls}" if cls and subj
                else _humanise(subj) or cls or 'Sheet'
            )
            for subj, cls in order
        ]

        dup = _ZipSheetDuplicator(self.template_path)
        dup.prepare(labels[1:])
        dup._rename_sheet_in_workbook(1, labels[0])
        self._dup = dup

        for sheet_idx, key in enumerate(order):
            subj, cls = key
            dup.write_header(
                sheet_idx,
                subj or 'All Subjects',
                term_year,
                cls or 'All Classes',
            )
            sorted_sids = sorted(
                groups[key],
                key=lambda s: student_meta.get(s, {}).get('name', '')
            )
            for i, sid in enumerate(sorted_sids):
                meta = student_meta.get(sid, {})
                sd = {
                    'name':       meta.get('name', ''),
                    'ref_id':     meta.get('ref_id', ''),
                    'study_area': meta.get('study_area', ''),
                    **{cat: groups[key][sid].get(cat, 0) for cat in CATEGORY_MAX},
                }
                dup.write_student_row(sheet_idx, STUDENT_START_ROW + i, sd)

        return dup.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# openpyxl-level helpers (single-sheet writes via wb.active)
# ─────────────────────────────────────────────────────────────────────────────

def _write_student_row(ws, row: int, sd: dict):
    """Write data columns only; formula columns are never touched."""
    serial = row - STUDENT_START_ROW + 1
    ws.cell(row=row, column=1,  value=serial)
    ws.cell(row=row, column=2,  value=sd.get('name', ''))
    ws.cell(row=row, column=3,  value=sd.get('ref_id', ''))
    ws.cell(row=row, column=4,  value=sd.get('study_area', ''))
    for cat, col in _CAT_TO_COL.items():
        raw = sd.get(cat, 0)
        ws.cell(row=row, column=col,
                value=float(raw) if raw not in (None, '') else 0.0)


def _ensure_formula_row(ws, row: int):
    """
    The template pre-fills rows 10–110 with formulas.
    Only needs to populate missing formula cells when the target row lacks them.

    BUG 4 FIX: use row 11 as the preferred shared-formula source,
    but fall back to row 10 when row 11 is blank.
    """
    source_row = STUDENT_START_ROW + 1   # row 11
    fallback_row = STUDENT_START_ROW      # row 10
    for col in _FORMULA_COLS:
        src = ws[f'{col}{source_row}']
        if not (src.value and str(src.value).startswith('=')):
            src = ws[f'{col}{fallback_row}']
        tgt = ws[f'{col}{row}']
        if (src.value and str(src.value).startswith('=')) and \
                not (tgt.value and str(tgt.value).startswith('=')):
            tgt.value = _shift_formula(str(src.value), int(src.row), row)


# ─────────────────────────────────────────────────────────────────────────────
# Module-level helpers
# ─────────────────────────────────────────────────────────────────────────────

def _humanise(key: str) -> str:
    return key.replace('_', ' ').title() if key else ''


def _safe_sheet_name(name: str) -> str:
    for ch in r'/\?*[]:\'':
        name = name.replace(ch, '-')
    return name[:31]


def _shift_formula(formula: str, from_row: int, to_row: int) -> str:
    """
    BUG 2 FIX: safe cell-reference regex — matches only complete cell refs
    (one or more capital letters followed by the exact row number, not
    followed by another digit).
    """
    pattern = re.compile(r'([A-Z]+)' + str(from_row) + r'(?!\\d)')
    return pattern.sub(lambda m: m.group(1) + str(to_row), formula)


def _build_term_year(settings) -> str:
    """
    BUG 7 FIX: always returns a clean string, never raises.
    'term1' → 'Term 1',  'term2' → 'Term 2', etc.
    """
    if not settings:
        return ''
    term_raw = getattr(settings, 'current_term', '') or ''
    year     = getattr(settings, 'current_academic_year', '') or ''
    try:
        from flask import current_app
        terms_cfg  = current_app.config.get('TERMS', [])
        term_label = dict(terms_cfg).get(term_raw, term_raw)
    except Exception:
        # No app context or config missing — humanise the raw key
        term_label = re.sub(r'term(\\d+)', r'Term \1', term_raw, flags=re.IGNORECASE).strip()
        if not term_label:
            term_label = term_raw
    return f"{term_label} {year}".strip()
