import os
import tempfile
from openpyxl import Workbook, load_workbook
from flask_login import login_user

from app import app, db, export_csv, export_student_csv, export_student_excel, download_template
from excel_utils import ExcelBulkImporter
from models import User, Student, Assessment, Setting
from template_updater import AssessmentTemplateUpdater, calculate_scores_from_template


def _create_minimal_school_template(path):
    wb = Workbook()
    ws = wb.active
    row = 10
    formulas = {
        'G': '=MIN(100,(SUM(E10:F10)))',
        'J': '=MIN(100,(SUM(H10:I10)))',
        'M': '=MIN(100,(SUM(K10:L10)))',
        'P': '=MIN(500,(SUM(G10,J10,M10,N10,O10)))',
        'Q': '=P10/500*100',
        'R': '=MIN(50,(ROUNDUP(Q10/2,0)))',
        'T': '=MIN(50,(ROUNDUP(S10/2,0)))',
        'U': '=MIN(100,(SUM(R10,T10)))',
        'V': '=U10/100',
        'W': '=IF(U10>=80,4,IF(U10>=70,3.5,IF(U10>=65,2.5,IF(U10>=60,2,IF(U10>=55,1.5,IF(U10>=50,1,IF(U10>=45,0.5,0)))))))',
        'X': '=IF(U10>=80,"A1",IF(U10>=70,"B2",IF(U10>=65,"B3",IF(U10>=60,"C4",IF(U10>=55,"C5",IF(U10>=50,"C6",IF(U10>=45,"D7",IF(U10>=40,"E8","F9"))))))))',
    }
    for col, formula in formulas.items():
        ws[f'{col}{row}'] = formula
    wb.save(path)


def test_calculate_scores_from_template_returns_expected_results():
    raw_scores = {
        'ica1': 40,
        'ica2': 45,
        'icp1': 30,
        'icp2': 35,
        'gp1': 20,
        'gp2': 25,
        'practical': 80,
        'mid_term': 90,
        'end_term': 85,
    }

    result = calculate_scores_from_template(raw_scores)

    assert result['ica_total'] == 85.0
    assert result['icp_total'] == 65.0
    assert result['gp_total'] == 45.0
    assert result['total_class_score'] == 365.0
    assert result['avg_class_score'] == 37.0
    assert result['avg_exam_score'] == 43.0
    assert result['final_score'] == 80.0
    assert result['gpa'] == 4.0
    assert result['grade'] == 'A1'


def test_assessment_template_updater_copies_row_formulas(tmp_path):
    template_path = tmp_path / 'student_template.xlsx'
    output_path = tmp_path / 'exported.xlsx'
    _create_minimal_school_template(str(template_path))

    updater = AssessmentTemplateUpdater(str(template_path))
    updater.load_template()
    updater.add_student(11, {
        'name': 'Test Student',
        'ref_id': 'REF001',
        'study_area': 'Mathematics',
        'ica1': 20,
        'ica2': 25,
        'icp1': 15,
        'icp2': 10,
        'gp1': 5,
        'gp2': 10,
        'practical': 30,
        'mid_term': 40,
        'end_term': 50,
    })
    updater.save_workbook(str(output_path))

    workbook = load_workbook(str(output_path), data_only=False)
    ws = workbook.active

    assert ws['E11'].value == 20.0
    assert ws['F11'].value == 25.0
    assert ws['G11'].value == '=MIN(100,(SUM(E11:F11)))'
    assert ws['P11'].value == '=MIN(500,(SUM(G11,J11,M11,N11,O11)))'
    assert ws['X11'].value == '=IF(U11>=80,"A1",IF(U11>=70,"B2",IF(U11>=65,"B3",IF(U11>=60,"C4",IF(U11>=55,"C5",IF(U11>=50,"C6",IF(U11>=45,"D7",IF(U11>=40,"E8","F9"))))))))'


def test_excel_bulk_importer_accepts_reference_number_header(tmp_path):
    path = tmp_path / 'assessment_import.xlsx'
    wb = Workbook()
    ws = wb.active
    for _ in range(8):
        ws.append([None] * 10)
    ws.append([
        'Student Number', 'Name of Students', 'Reference Number', 'Learning Area',
        'category', 'subject', 'score', 'max_score', 'term', 'session',
        'assessor', 'comments'
    ])
    ws.append([
        'STU001', 'Jane Doe', 'REF001', 'Mathematics',
        'ica1', 'Mathematics', 40, 50, 'First Term', '2024/2025',
        'Teacher A', 'Good effort'
    ])
    wb.save(str(path))

    imported = ExcelBulkImporter(str(path)).import_assessments()
    assert len(imported) == 1
    assert imported[0]['student_number'] == 'STU001'
    assert imported[0]['reference_number'] == 'REF001'
    assert imported[0]['category'] == 'ica1'
    assert imported[0]['score'] == 40


def _setup_db_with_template(app, tmp_path):
    temp_db = tmp_path / 'test.db'
    app.config['TESTING'] = True
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{temp_db}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['TEMPLATE_FOLDER'] = str(tmp_path)
    app.config['UPLOAD_FOLDER'] = str(tmp_path)
    os.makedirs(app.config['TEMPLATE_FOLDER'], exist_ok=True)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    with app.app_context():
        ext = app.extensions.get('sqlalchemy')
        if ext and hasattr(ext, '_app_engines'):
            ext._app_engines[app].clear()
            options = {'url': app.config['SQLALCHEMY_DATABASE_URI'], **ext._engine_options}
            engine = ext._make_engine(None, options, app)
            ext._app_engines[app][None] = engine
        db.create_all()
        setting = Setting(current_term='First Term', current_academic_year='2024-2025')
        db.session.add(setting)
        db.session.commit()

    return temp_db


def test_export_student_excel_route_creates_file(tmp_path):
    template_path = tmp_path / 'student_template.xlsx'
    _create_minimal_school_template(str(template_path))

    _setup_db_with_template(app, tmp_path)

    with app.app_context():
        admin_user = User(username='admin_test', password_hash='x', role='admin')
        db.session.add(admin_user)
        db.session.commit()

        student = Student(
            first_name='Jane',
            last_name='Doe',
            student_number='STU100',
            class_name='form1',
            study_area='mathematics'
        )
        db.session.add(student)
        db.session.commit()

        assessment = Assessment(
            student_id=student.id,
            category='ica1',
            subject='mathematics',
            class_name='form1',
            score=40.0,
            max_score=50.0,
            teacher_id=admin_user.id,
        )
        db.session.add(assessment)
        db.session.commit()

        with app.test_request_context():
            login_user(admin_user)
            response = export_student_excel(student.id)

        assert response.status_code == 200
        assert 'attachment' in response.headers.get('Content-Disposition', '')


def test_export_csv_route_redirects_to_excel(tmp_path):
    template_path = tmp_path / 'student_template.xlsx'
    _create_minimal_school_template(str(template_path))

    _setup_db_with_template(app, tmp_path)

    with app.app_context():
        admin_user = User(username='admin_test', password_hash='x', role='admin')
        db.session.add(admin_user)
        db.session.commit()

        with app.test_request_context():
            login_user(admin_user)
            response = export_csv()

        assert response.status_code == 302
        assert response.headers.get('Location', '').endswith('/export/assessments/excel')


def test_export_student_csv_route_redirects_to_excel(tmp_path):
    template_path = tmp_path / 'student_template.xlsx'
    _create_minimal_school_template(str(template_path))

    _setup_db_with_template(app, tmp_path)

    with app.app_context():
        admin_user = User(username='admin_test2', password_hash='x', role='admin')
        db.session.add(admin_user)
        db.session.commit()

        student = Student(
            first_name='Jane',
            last_name='Doe',
            student_number='STU101',
            class_name='form1',
            study_area='mathematics'
        )
        db.session.add(student)
        db.session.commit()

        with app.test_request_context():
            login_user(admin_user)
            response = export_student_csv(student.id)

        assert response.status_code == 302
        assert response.headers.get('Location', '').endswith(f'/export/excel/student/{student.id}')


def test_download_template_student_route_serves_template(tmp_path):
    template_path = tmp_path / 'student_template.xlsx'
    _create_minimal_school_template(str(template_path))

    _setup_db_with_template(app, tmp_path)

    with app.app_context():
        admin_user = User(username='admin_download', password_hash='x', role='admin')
        db.session.add(admin_user)
        db.session.commit()

        with app.test_request_context():
            login_user(admin_user)
            response = download_template('student')

        assert response.status_code == 200
        assert 'attachment' in response.headers.get('Content-Disposition', '')
