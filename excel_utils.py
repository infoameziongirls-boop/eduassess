"""
Excel utility classes for EduAssess
"""
import os
import tempfile
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Protection
from datetime import datetime


class ExcelTemplateHandler:
    def __init__(self, template_path):
        self.template_path = template_path

    def load_template(self):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found: {self.template_path}")
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        shutil.copy2(self.template_path, temp_path)
        self.temp_path = temp_path
        return load_workbook(temp_path)

    def export_student_to_template(self, student, assessments, output_path, config):
        wb = self.load_template()
        ws = wb.active
        self._write_student_info(ws, student)
        self._write_assessments(ws, assessments, config)
        self._write_summary(ws, assessments, config)
        wb.save(output_path)
        return output_path

    def _write_student_info(self, ws, student):
        ws['B2'] = student.student_number
        ws['B3'] = student.full_name()
        ws['B4'] = student.study_area if student.study_area else ""
        ws['B5'] = datetime.now().strftime('%Y-%m-%d')

    def _write_assessments(self, ws, assessments, config):
        by_category = {}
        for assessment in assessments:
            if assessment.category not in by_category:
                by_category[assessment.category] = []
            by_category[assessment.category].append(assessment)

        category_rows = {
            "IA": 8, "IPA": 13, "PP": 18, "MSE": 23, "ETE": 28
        }
        for category, start_row in category_rows.items():
            if category in by_category:
                for idx, assessment in enumerate(by_category[category]):
                    row = start_row + idx
                    ws.cell(row=row, column=1, value=assessment.subject or "")
                    ws.cell(row=row, column=2, value=assessment.score)
                    ws.cell(row=row, column=3, value=assessment.max_score)
                    ws.cell(row=row, column=4, value=assessment.get_percentage())
                    ws.cell(row=row, column=5, value=assessment.term or "")

    def _write_summary(self, ws, assessments, config):
        from collections import defaultdict
        summary = defaultdict(lambda: {"total_score": 0, "total_max": 0, "count": 0})
        for assessment in assessments:
            cat = assessment.category
            summary[cat]["total_score"] += assessment.score
            summary[cat]["total_max"] += assessment.max_score
            summary[cat]["count"] += 1

        summary_row = 35
        for idx, (cat, label) in enumerate(config['CATEGORY_LABELS'].items()):
            row = summary_row + idx
            if cat in summary and summary[cat]["total_max"] > 0:
                avg = (summary[cat]["total_score"] / summary[cat]["total_max"]) * 100
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=summary[cat]["count"])
                ws.cell(row=row, column=3, value=f"{avg:.2f}%")
            else:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=0)
                ws.cell(row=row, column=3, value="N/A")


class ExcelBulkImporter:
    def __init__(self, file_path):
        self.file_path = file_path

    def import_assessments(self, start_row=2):
        wb = load_workbook(self.file_path, data_only=True)
        try:
            ws = wb.active

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                return str(value).strip()

            def normalize_header(value):
                if value is None:
                    return ''
                return str(value).strip().lower().replace('.', '').replace('_', ' ')

            header_row = None
            header_map = {}
            for idx, row in enumerate(ws.iter_rows(min_row=1,
                                                 max_row=min(15, ws.max_row),
                                                 values_only=True), start=1):
                if not any(row):
                    continue
                normalized = [normalize_header(cell) for cell in row]
                if any(label in normalized for label in (
                        'student number', 'reference number', 'ref id',
                        'category', 'subject', 'score', 'term', 'session')):
                    header_row = idx
                    for col_idx, label in enumerate(normalized):
                        if label in ('student number', 'student number or reference number',
                                     'student no', 'student id'):
                            header_map['student_number'] = col_idx
                        elif label in ('reference number', 'reference no', 'ref id'):
                            header_map['reference_number'] = col_idx
                        elif label == 'category':
                            header_map['category'] = col_idx
                        elif label == 'subject':
                            header_map['subject'] = col_idx
                        elif label == 'score':
                            header_map['score'] = col_idx
                        elif label in ('max score', 'maximum score', 'max_score'):
                            header_map['max_score'] = col_idx
                        elif label == 'term':
                            header_map['term'] = col_idx
                        elif label in ('academic year', 'academic_year'):
                            header_map['academic_year'] = col_idx
                        elif label == 'session':
                            header_map['session'] = col_idx
                        elif label == 'assessor':
                            header_map['assessor'] = col_idx
                        elif label == 'comments':
                            header_map['comments'] = col_idx
                    break

            data_row_start = header_row + 1 if header_row else start_row
            assessments = []
            for row in ws.iter_rows(min_row=data_row_start, values_only=True):
                if not any(row):
                    continue

                if header_row:
                    student_number = normalize(row[header_map['student_number']]) \
                        if 'student_number' in header_map and len(row) > header_map['student_number'] else None
                    reference_number = normalize(row[header_map['reference_number']]) \
                        if 'reference_number' in header_map and len(row) > header_map['reference_number'] else None
                    category = normalize(row[header_map['category']]) \
                        if 'category' in header_map and len(row) > header_map['category'] else None
                    subject = normalize(row[header_map['subject']]) \
                        if 'subject' in header_map and len(row) > header_map['subject'] else None
                    score = row[header_map['score']] if 'score' in header_map and len(row) > header_map['score'] else None
                    max_score = row[header_map['max_score']] if 'max_score' in header_map and len(row) > header_map['max_score'] else None
                    term = normalize(row[header_map['term']]) \
                        if 'term' in header_map and len(row) > header_map['term'] else None
                    session = normalize(row[header_map['session']]) \
                        if 'session' in header_map and len(row) > header_map['session'] else None
                    academic_year = normalize(row[header_map['academic_year']]) \
                        if 'academic_year' in header_map and len(row) > header_map['academic_year'] else None
                    assessor = normalize(row[header_map['assessor']]) \
                        if 'assessor' in header_map and len(row) > header_map['assessor'] else None
                    comments = normalize(row[header_map['comments']]) \
                        if 'comments' in header_map and len(row) > header_map['comments'] else ""
                else:
                    student_number = normalize(row[0]) if len(row) > 0 else None
                    reference_number = normalize(row[2]) if len(row) > 2 else None
                    category = normalize(row[1]) if len(row) > 1 else None
                    subject = normalize(row[2]) if len(row) > 2 else None
                    score = row[3] if len(row) > 3 else None
                    max_score = row[4] if len(row) > 4 else None
                    term = normalize(row[5]) if len(row) > 5 else None
                    session = normalize(row[6]) if len(row) > 6 else None
                    academic_year = normalize(row[7]) if len(row) > 7 else None
                    assessor = normalize(row[8]) if len(row) > 8 else None
                    comments = normalize(row[9]) if len(row) > 9 else ""

                assessment_data = {
                    'student_number': student_number,
                    'reference_number': reference_number,
                    'category': category,
                    'subject': subject,
                    'score': score,
                    'max_score': max_score,
                    'term': term,
                    'session': session,
                    'academic_year': academic_year,
                    'assessor': assessor,
                    'comments': comments,
                }
                if (assessment_data['student_number'] or assessment_data['reference_number']) and \
                        assessment_data['score'] is not None:
                    assessments.append(assessment_data)
            return assessments
        finally:
            wb.close()


class ClassScoreSheetImporter:
    """
    Imports a 'wide format' class scoresheet: ONE ROW PER STUDENT with all
    assessment categories (ica1, ica2, icp1, icp2, gp1, gp2, practical,
    mid_term, end_term) as columns, instead of one row per category like
    ExcelBulkImporter expects. This is what lets a teacher fill in scores
    for an entire class, every category, in a single spreadsheet.
    """

    CATEGORY_HEADER_ALIASES = {
        'ica1': ('ica1', 'ica 1', 'individual class assessment 1', 'individual assessment 1'),
        'ica2': ('ica2', 'ica 2', 'individual class assessment 2', 'individual assessment 2'),
        'icp1': ('icp1', 'icp 1', 'individual class project 1'),
        'icp2': ('icp2', 'icp 2', 'individual class project 2'),
        'gp1': ('gp1', 'gp 1', 'group project/research 1', 'group project 1'),
        'gp2': ('gp2', 'gp 2', 'group project/research 2', 'group project 2'),
        'practical': ('practical', 'practical portfolio'),
        'mid_term': ('mid term', 'mid-term', 'mid semester exam', 'mid-semester exam', 'mid term exam', 'mid-term exam'),
        'end_term': ('end term', 'end of term', 'end term exam', 'end of term exam'),
    }

    # Order used when a category can't be matched by header text (fallback
    # positional layout produced by create_class_scoresheet_template()).
    CATEGORY_ORDER = ['ica1', 'ica2', 'icp1', 'icp2', 'gp1', 'gp2',
                      'practical', 'mid_term', 'end_term']

    def __init__(self, file_path):
        self.file_path = file_path

    def import_scoresheet(self, start_row=2):
        """
        Returns a list of dicts:
            {
              'system_id': int or None,        # authoritative match, from the
                                                # hidden/protected internal ID
                                                # column — teachers cannot edit
                                                # or forge this in Excel
              'student_number': str or None,   # display-only, NOT used to
                                                # resolve identity if system_id
                                                # is present
              'reference_number': str or None, # same — display-only
              'name': str or None,
              'scores': {category: float, ...}   # only categories with a value
            }
        """
        wb = load_workbook(self.file_path, data_only=True)
        try:
            ws = wb.active

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                return str(value).strip()

            def normalize_header(value):
                if value is None:
                    return ''
                return str(value).strip().lower().replace('.', '').replace('_', ' ')

            header_row = None
            header_map = {}
            category_cols = {}
            for idx, row in enumerate(ws.iter_rows(min_row=1,
                                                 max_row=min(15, ws.max_row),
                                                 values_only=True), start=1):
                if not any(row):
                    continue
                normalized = [normalize_header(cell) for cell in row]
                if any(label in normalized for label in (
                        'student number', 'reference number', 'ref id', 'ref no')):
                    header_row = idx
                    for col_idx, label in enumerate(normalized):
                        if label in ('student number', 'student no'):
                            header_map['student_number'] = col_idx
                        elif label in ('reference number', 'reference no', 'ref id', 'ref no'):
                            header_map['reference_number'] = col_idx
                        elif label in ('name of student', 'name of students', 'name', 'student name', 'full name'):
                            header_map['name'] = col_idx
                        elif label in ('system id', 'record id', 'db id', 'student db id',
                                       'internal id'):
                            header_map['system_id'] = col_idx
                        else:
                            for cat, aliases in self.CATEGORY_HEADER_ALIASES.items():
                                if label in aliases:
                                    category_cols[cat] = col_idx
                                    break
                    break

            if header_row is None:
                raise ValueError(
                    'Could not find a header row. The scoresheet must include '
                    'a "Student Number" or "Reference Number" column.'
                )

            if not category_cols:
                raise ValueError(
                    'No recognizable category columns found (ICA1, ICA2, ICP1, '
                    'ICP2, GP1, GP2, Practical, Mid Term, End Term).'
                )

            students = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(row):
                    continue

                def cell(col_idx):
                    return row[col_idx] if col_idx is not None and len(row) > col_idx else None

                system_id_raw = cell(header_map.get('system_id'))
                system_id = None
                if system_id_raw is not None and str(system_id_raw).strip() != '':
                    try:
                        system_id = int(float(system_id_raw))
                    except (TypeError, ValueError):
                        system_id = None

                student_number = normalize(cell(header_map.get('student_number')))
                reference_number = normalize(cell(header_map.get('reference_number')))
                name = normalize(cell(header_map.get('name')))

                if not system_id and not student_number and not reference_number:
                    continue

                scores = {}
                for cat, col_idx in category_cols.items():
                    raw_val = cell(col_idx)
                    if raw_val is None or (isinstance(raw_val, str) and not raw_val.strip()):
                        continue
                    try:
                        scores[cat] = float(raw_val)
                    except (TypeError, ValueError):
                        continue

                if not scores:
                    continue

                students.append({
                    'system_id': system_id,
                    'student_number': student_number,
                    'reference_number': reference_number,
                    'name': name,
                    'scores': scores,
                })
            return students
        finally:
            wb.close()


class StudentBulkImporter:
    def __init__(self, file_path):
        self.file_path = file_path

    def import_students(self, start_row=2):
        wb = load_workbook(self.file_path, data_only=True)
        try:
            ws = wb.active
            students = []

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                return str(value).strip()

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not any(row):
                    continue
                student_data = {
                    'student_number': normalize(row[0]) if len(row) > 0 else None,
                    'first_name':     normalize(row[1]) if len(row) > 1 else None,
                    'last_name':      normalize(row[2]) if len(row) > 2 else None,
                    'middle_name':    normalize(row[3]) if len(row) > 3 else None,
                    'class_name':     normalize(row[4]) if len(row) > 4 else None,
                    'study_area':     normalize(row[5]) if len(row) > 5 else None,
                }
                if student_data['student_number'] and student_data['first_name'] and student_data['last_name']:
                    students.append(student_data)
            return students
        finally:
            wb.close()


class TeacherBulkImporter:
    def __init__(self, file_path):
        self.file_path = file_path

    def import_teachers(self, start_row=2):
        wb = load_workbook(self.file_path, data_only=True)
        try:
            ws = wb.active
            teachers = []

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                return str(value).strip()

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not any(row):
                    continue
                teacher_data = {
                    'username': normalize(row[0]) if len(row) > 0 else None,
                    'password': normalize(row[1]) if len(row) > 1 else None,
                    'role':     normalize(row[2]) if len(row) > 2 else None,
                    'subject':  normalize(row[3]) if len(row) > 3 else None,
                    'classes':  normalize(row[4]) if len(row) > 4 else None,
                }
                if teacher_data['username']:
                    teachers.append(teacher_data)
            return teachers
        finally:
            wb.close()


class QuestionBulkImporter:
    def __init__(self, file_path):
        self.file_path = file_path

    def import_questions(self, start_row=2):
        wb = load_workbook(self.file_path, data_only=True)
        try:
            ws = wb.active
            questions = []
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not any(row):
                    continue
                question_data = {
                    'question_text':  row[0],
                    'question_type':  row[1],
                    'option_a':       row[2] if len(row) > 2 else None,
                    'option_b':       row[3] if len(row) > 3 else None,
                    'option_c':       row[4] if len(row) > 4 else None,
                    'option_d':       row[5] if len(row) > 5 else None,
                    'correct_answer': row[6],
                    'difficulty':     row[7] if len(row) > 7 else 'medium',
                    'explanation':    row[8] if len(row) > 8 else None,
                }
                if question_data['question_text'] and question_data['question_type'] and question_data['correct_answer']:
                    if question_data['question_type'].lower() == 'mcq':
                        question_data['options'] = [
                            question_data['option_a'], question_data['option_b'],
                            question_data['option_c'], question_data['option_d']
                        ]
                    else:
                        question_data['options'] = None
                    questions.append(question_data)
            return questions
        finally:
            wb.close()


def create_default_template(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ASSESSMENT TEMPLATE"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    ws['A1'] = "SCHOOL:"
    ws['A2'] = "SUBJECT:"
    ws['A3'] = "TERM/YEAR:"
    ws['A4'] = "FORM:"

    headers = [
        "Serial Number", "Name of Students", "Ref. Id", "Study Area",
        "ICA1", "ICA2", "SUB TOTAL (I.C.A.)", "ICP1", "ICP2", "SUB TOTAL TEST(C.P)",
        "GP1", "GP2", "SUB TOTAL (G.P)", "Practical", "Mid Term", "Total Class",
        "%", "AVG. CLASS", "End Term", "AVG. EXAMS SC.", "Total 50 + 50",
        "GPA", "Grade", None, "INSTRUCTIONS"
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=9, column=idx, value=header)
        if header:
            ws.cell(row=9, column=idx).font = header_font
            ws.cell(row=9, column=idx).fill = header_fill

    row = 10
    ws[f"G{row}"] = f"=MIN(100,(SUM(E{row}:F{row})))"
    ws[f"J{row}"] = f"=MIN(100,(SUM(H{row}:I{row})))"
    ws[f"M{row}"] = f"=MIN(100,(SUM(K{row}:L{row})))"
    ws[f"P{row}"] = f"=MIN(500,(SUM(G{row},J{row},M{row},N{row},O{row})))"
    ws[f"Q{row}"] = f"=P{row}/500*100"
    ws[f"R{row}"] = f"=MIN(50,(ROUNDUP(SUM(Q{row})/2,0)))"
    ws[f"T{row}"] = f"=MIN(50,(ROUNDUP(SUM(S{row})/2,0)))"
    ws[f"U{row}"] = f"=MIN(100,(SUM(R{row},T{row})))"
    ws[f"V{row}"] = f"=U{row}"
    ws[f"W{row}"] = (
        f'=IF(U{row}>=80,"4.0",IF(U{row}>=70,"3.5",IF(U{row}>=65,"3.0",'
        f'IF(U{row}>=60,"2.5",IF(U{row}>=55,"2.0",IF(U{row}>=50,"1.5",'
        f'IF(U{row}>=45,"1.0",IF(U{row}>=40,"0.5","0.0"))))))))'
    )
    ws[f"X{row}"] = (
        f'=IF(U{row}>=80,"A1",IF(U{row}>=70,"B2",IF(U{row}>=65,"B3",'
        f'IF(U{row}>=60,"C4",IF(U{row}>=55,"C5",IF(U{row}>=50,"C6",'
        f'IF(U{row}>=45,"D7",IF(U{row}>=40,"E8","F9"))))))))'
    )

    for col, width in zip(
        'ABCDEFGHIJKLMNOPQRSTUVWX',
        [10, 25, 15, 15, 8, 8, 15, 8, 8, 15, 8, 8, 15, 12, 12, 15, 10, 12, 12, 15, 10, 10, 12, 20]
    ):
        ws.column_dimensions[col].width = width

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path


def create_class_scoresheet_template(output_path, students=None, subject_label=None,
                                     class_label=None, category_labels=None):
    """
    Builds the wide-format 'class scoresheet' workbook: one row per student,
    one column per assessment category. If `students` is provided (a list of
    Student model objects or dicts with id/student_number/reference_number/
    name/study_area), the roster is pre-filled so the teacher only has to
    type in scores — nothing else.

    Identity protection: when a roster is supplied, each student's database
    ID is written into a hidden column, and the sheet is protected so that
    only the score columns can be edited. Import always resolves identity
    from that hidden ID column, never from the (locked, but still
    tamper-resistant-in-depth) Student Number / Name / Reference Number
    cells — so teachers cannot add students, rename them, or repoint a row
    at a different student by editing those cells.

    category_labels: optional dict of {category_key: column_header}. Falls
    back to the standard 9 categories if not supplied.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Scoresheet"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.password = 'eduassess'

    if subject_label or class_label:
        ws['A1'] = "SUBJECT:"
        ws['B1'] = subject_label or ""
        ws['A2'] = "CLASS:"
        ws['B2'] = class_label or ""

    default_categories = [
        ('ica1', 'ICA1'), ('ica2', 'ICA2'),
        ('icp1', 'ICP1'), ('icp2', 'ICP2'),
        ('gp1', 'GP1'), ('gp2', 'GP2'),
        ('practical', 'Practical'), ('mid_term', 'Mid Term'), ('end_term', 'End Term'),
    ]
    categories = [(key, (category_labels or {}).get(key, label))
                  for key, label in default_categories]

    # Fixed column layout. "System ID" is the hidden identity column the
    # importer trusts first; it is deliberately placed beyond the visible
    # roster columns and then hidden so teachers cannot reliably alter it.
    ID_COL = 4 + len(categories) + 1  # after Student Number/Name/Ref/Study Area + categories

    header_row = 4
    visible_headers = ['Student Number', 'Name of Student', 'Reference Number', 'Study Area'] + \
                      [label for _, label in categories]
    for idx, header in enumerate(visible_headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    id_header_cell = ws.cell(row=header_row, column=ID_COL, value='System ID')
    id_header_cell.font = header_font
    id_header_cell.fill = header_fill
    ws.column_dimensions[get_column_letter_local(ID_COL)].hidden = True

    def _get(obj, attr):
        if obj is None:
            return ''
        if isinstance(obj, dict):
            return obj.get(attr, '') or ''
        return getattr(obj, attr, '') or ''

    def _display_name(obj):
        if isinstance(obj, dict):
            if obj.get('name'):
                return obj['name']
            parts = [obj.get('first_name', ''), obj.get('middle_name', ''), obj.get('last_name', '')]
            return ' '.join(p for p in parts if p).strip()
        if hasattr(obj, 'full_name'):
            return obj.full_name()
        return ''

    NAME_NUMBER_COLS = (1, 2, 3, 4)  # Student Number, Name, Reference Number, Study Area
    CATEGORY_COLS = range(5, 5 + len(categories))

    row_num = header_row + 1
    if students:
        for student in students:
            student_id = _get(student, 'id') or _get(student, 'student_id')
            ws.cell(row=row_num, column=1, value=_get(student, 'student_number'))
            ws.cell(row=row_num, column=2, value=_display_name(student))
            ws.cell(row=row_num, column=3, value=_get(student, 'reference_number'))
            ws.cell(row=row_num, column=4, value=_get(student, 'study_area'))
            ws.cell(row=row_num, column=ID_COL, value=student_id)

            for col in NAME_NUMBER_COLS:
                ws.cell(row=row_num, column=col).protection = locked
            ws.cell(row=row_num, column=ID_COL).protection = locked
            for col in CATEGORY_COLS:
                ws.cell(row=row_num, column=col).protection = unlocked
            row_num += 1
    else:
        # Blank template with a couple of sample rows for reference.
        sample_rows = [
            ["STU001", "John Doe", "REF001", "Mathematics"],
            ["STU002", "Jane Smith", "REF002", "Mathematics"],
        ]
        for sample in sample_rows:
            for col_idx, value in enumerate(sample, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)
            ws.cell(row=row_num, column=ID_COL, value='')
            for col in NAME_NUMBER_COLS:
                ws.cell(row=row_num, column=col).protection = locked
            ws.cell(row=row_num, column=ID_COL).protection = locked
            for col in CATEGORY_COLS:
                ws.cell(row=row_num, column=col).protection = unlocked
            row_num += 1

    widths = [15, 25, 15, 20] + [10] * len(categories)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter_local(idx)].width = width

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path


def get_column_letter_local(idx):
    """Small local helper so this module doesn't need an extra top-level import
    if openpyxl.utils isn't already imported elsewhere in this file."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def create_student_import_template(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    headers = ["Student Number", "First Name", "Last Name", "Middle Name", "Class", "Study Area"]
    for idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=idx + 1, value=header)
        cell.font = header_font
        cell.fill = header_fill

    sample_data = [
        ["STU001", "John", "Doe", "Michael", "Form 1", "Home Economics A"],
        ["STU002", "Jane", "Smith", "", "Form 2", "General Arts 4B"],
        ["STU003", "Bob", "Johnson", "William", "Form 3", "Business A"],
    ]
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=row_idx, column=col_idx + 1, value=value)

    for idx, width in enumerate([15, 15, 15, 15, 10, 20]):
        ws.column_dimensions[chr(65 + idx)].width = width

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path


def create_teacher_import_template(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Import"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    headers = ["Username", "Password", "Role", "Subject", "Classes"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    sample_data = [
        ["teacher1", "Teacher@123", "teacher", "Mathematics", "Form 1"],
        ["teacher2", "Teacher@123", "teacher", "English Language", "Form 2"],
        ["admin1",   "Admin@123",   "admin",   "",               ""]
    ]
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for idx, width in enumerate([20, 20, 15, 20, 25], start=1):
        ws.column_dimensions[chr(64+idx)].width = width
    wb.save(output_path)
    return output_path


def create_question_import_template(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Import"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    headers = ["Question Text", "Question Type", "Option A", "Option B",
               "Option C", "Option D", "Correct Answer", "Difficulty", "Explanation"]
    for idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=idx+1, value=header)
        cell.font = header_font
        cell.fill = header_fill
    sample_data = [
        ["What is the capital of France?", "mcq", "Paris", "London", "Berlin", "Madrid", "A", "easy", "Paris is the capital of France."],
        ["The Earth is round.", "true_false", "", "", "", "", "True", "easy", "Scientific fact."],
        ["What is 2 + 2?", "short_answer", "", "", "", "", "4", "easy", "Basic arithmetic."]
    ]
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=row_idx, column=col_idx+1, value=value)
    for idx, width in enumerate([40, 15, 15, 15, 15, 15, 15, 10, 30]):
        ws.column_dimensions[chr(65+idx)].width = width
    wb.save(output_path)
    return output_path
